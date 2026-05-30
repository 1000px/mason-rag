import json
import re
import sqlite3
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from langchain_core.documents import Document
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from src.config.settings import resolve_path


class TableStore:
    """SQLite-based structured storage for Excel data. One table per sheet."""

    def __init__(self, db_path: str | Path | None = None):
        if db_path is None:
            db_path = resolve_path("data/tables.db")
        else:
            db_path = Path(db_path)
        db_path.parent.mkdir(parents=True, exist_ok=True)
        self.db_path = str(db_path)
        self._init_meta()

    def _connect(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL")
        return conn

    def _init_meta(self):
        with self._connect() as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS _table_meta (
                    table_name TEXT PRIMARY KEY,
                    file_name TEXT NOT NULL,
                    sheet_name TEXT NOT NULL,
                    row_count INTEGER DEFAULT 0,
                    columns TEXT NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS _relationships (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    table_a TEXT NOT NULL,
                    column_a TEXT NOT NULL,
                    table_b TEXT NOT NULL,
                    column_b TEXT NOT NULL,
                    confidence REAL NOT NULL DEFAULT 0,
                    relation_type TEXT DEFAULT 'value_overlap',
                    row_count_a INTEGER,
                    row_count_b INTEGER,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)

    @staticmethod
    def _sanitize_name(name: str) -> str:
        name = re.sub(r"[^\w]", "_", name)
        name = re.sub(r"_+", "_", name).strip("_")
        if not name:
            name = "sheet"
        if name[0].isdigit():
            name = "_" + name
        return name

    def ingest_excel(self, file_path: Path) -> Dict[str, Any]:
        file_name = file_path.name
        ext = file_path.suffix.lower()

        if ext == ".xlsx":
            wb = load_workbook(str(file_path), data_only=True)
            sheets_data = {}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                headers = None
                for row in ws.iter_rows(values_only=True):
                    row = [str(c) if c is not None else "" for c in row]
                    if not any(row):
                        continue
                    if headers is None:
                        headers = row
                    else:
                        rows.append(row)
                if headers:
                    sheets_data[sheet_name] = (headers, rows)
        elif ext == ".xls":
            import xlrd
            wb = xlrd.open_workbook(str(file_path))
            sheets_data = {}
            for sheet in wb.sheets():
                headers = None
                rows = []
                for row_idx in range(sheet.nrows):
                    row = [str(sheet.cell_value(row_idx, col_idx)) for col_idx in range(sheet.ncols)]
                    if not any(row):
                        continue
                    if headers is None:
                        headers = row
                    else:
                        rows.append(row)
                if headers:
                    sheets_data[sheet.name] = (headers, rows)
        else:
            raise ValueError(f"不支持的格式: {ext}")

        result = {"file_name": file_name, "sheets": {}}

        with self._connect() as conn:
            for sheet_name, (headers, rows) in sheets_data.items():
                table_name = self._sanitize_name(f"{Path(file_name).stem}_{sheet_name}")

                safe_headers = []
                for h in headers:
                    sh = self._sanitize_name(h)
                    if not sh:
                        sh = "col"
                    while sh in safe_headers:
                        sh = sh + "_2"
                    safe_headers.append(sh)

                col_defs = ", ".join([f'"{h}" TEXT' for h in safe_headers])
                conn.execute(f'CREATE TABLE IF NOT EXISTS "{table_name}" ({col_defs})')
                conn.execute(f'DELETE FROM "{table_name}"')

                placeholders = ", ".join(["?" for _ in safe_headers])
                conn.executemany(
                    f'INSERT INTO "{table_name}" ({", ".join([f"\"{h}\"" for h in safe_headers])}) VALUES ({placeholders})',
                    [row[:len(safe_headers)] for row in rows],
                )

                conn.execute(
                    "INSERT OR REPLACE INTO _table_meta (table_name, file_name, sheet_name, row_count, columns) VALUES (?, ?, ?, ?, ?)",
                    (table_name, file_name, sheet_name, len(rows), json.dumps(safe_headers, ensure_ascii=False)),
                )

                result["sheets"][sheet_name] = {
                    "table_name": table_name,
                    "row_count": len(rows),
                    "columns": safe_headers,
                }

        return result

    def get_schema_for_llm(self) -> str:
        with self._connect() as conn:
            metas = conn.execute("SELECT table_name, sheet_name, row_count, columns FROM _table_meta").fetchall()

        if not metas:
            return "（当前无可用的结构化数据表）"

        relationships = self.get_relationships()

        lines = ["# 可用的结构化数据表\n"]
        for meta in metas:
            cols = json.loads(meta["columns"]) if isinstance(meta["columns"], str) else meta["columns"]
            lines.append(f"## 表: \"{meta['table_name']}\" (来源Sheet: {meta['sheet_name']}, 共 {meta['row_count']} 行)")
            lines.append(f"列: {', '.join(cols)}")
            sample = self.sample_rows(meta["table_name"], 3)
            if sample:
                lines.append("数据样例:")
                for row in sample:
                    lines.append(f"  {dict(row)}")
            lines.append("")

        if relationships:
            same_struct_rels = [r for r in relationships if r["relation_type"] == "same_column"]
            diff_struct_rels = [r for r in relationships if r["relation_type"] != "same_column"]

            lines.append("## 🔗 检测到的表间关联关系（可用于 JOIN）\n")

            if same_struct_rels:
                same_tables = set()
                for r in same_struct_rels[:5]:
                    same_tables.add(r["table_a"])
                    same_tables.add(r["table_b"])
                lines.append(
                    f"以下表结构相同（同名列 > 5 个），**禁止 INNER JOIN**，对比查询请用 **UNION ALL + GROUP BY**：\n"
                    f"  {', '.join(f'\"{t}\"' for t in same_tables)}\n"
                )

            if diff_struct_rels:
                lines.append("以下表可通过指定列进行 INNER JOIN 关联查询：")
                for rel in diff_struct_rels[:10]:
                    lines.append(
                        f"- \"{rel['table_a']}\".\"{rel['column_a']}\" ↔ \"{rel['table_b']}\".\"{rel['column_b']}\" "
                        f"(置信度: {rel['confidence']}, 类型: {rel['relation_type']})"
                    )

            lines.append("")
            lines.append("**跨表查询规则**：")
            lines.append("- 两张表结构相同 → 用 UNION ALL + GROUP BY 聚合，**禁止 INNER JOIN**")
            lines.append("- 两张表结构不同但有同名列 → 用 INNER JOIN 关联查询")
            lines.append("- UNION ALL 示例：SELECT 列, SUM(数量) FROM (SELECT 列, 数量 FROM 表1 UNION ALL SELECT 列, 数量 FROM 表2) GROUP BY 列")
            lines.append("")

        return "\n".join(lines)

    def sample_rows(self, table_name: str, n: int = 3) -> List[Dict]:
        with self._connect() as conn:
            rows = conn.execute(f'SELECT * FROM "{table_name}" LIMIT ?', (n,)).fetchall()
        return [dict(r) for r in rows]

    def execute_sql(self, sql: str) -> Tuple[List[Dict], List[str]]:
        with self._connect() as conn:
            cursor = conn.execute(sql)
            columns = [d[0] for d in cursor.description] if cursor.description else []
            rows = [dict(r) for r in cursor.fetchall()]
        return rows, columns

    def delete_by_file(self, file_name: str) -> int:
        with self._connect() as conn:
            metas = conn.execute(
                "SELECT table_name FROM _table_meta WHERE file_name = ?", (file_name,)
            ).fetchall()
            count = 0
            for meta in metas:
                conn.execute(f'DROP TABLE IF EXISTS "{meta["table_name"]}"')
                count += 1
            conn.execute("DELETE FROM _table_meta WHERE file_name = ?", (file_name,))
        return count

    def list_files(self) -> List[str]:
        with self._connect() as conn:
            rows = conn.execute("SELECT DISTINCT file_name FROM _table_meta").fetchall()
        return [r["file_name"] for r in rows]

    def get_table_names(self) -> List[str]:
        with self._connect() as conn:
            rows = conn.execute("SELECT table_name FROM _table_meta").fetchall()
        return [r["table_name"] for r in rows]

    def get_all_columns(self) -> Dict[str, List[str]]:
        with self._connect() as conn:
            metas = conn.execute("SELECT table_name, columns FROM _table_meta").fetchall()
        return {m["table_name"]: json.loads(m["columns"]) for m in metas}

    def _sample_column_values(self, table_name: str, column: str, n: int = 200) -> List[str]:
        with self._connect() as conn:
            rows = conn.execute(f'SELECT DISTINCT "{column}" FROM "{table_name}" WHERE "{column}" IS NOT NULL AND "{column}" != "" LIMIT ?', (n,)).fetchall()
        return [r[0] for r in rows if r[0]]

    @staticmethod
    def _column_value_overlap(values_a: List[str], values_b: List[str]) -> float:
        set_a = set(str(v).strip() for v in values_a)
        set_b = set(str(v).strip() for v in values_b)
        if not set_a or not set_b:
            return 0.0
        intersection = set_a & set_b
        union = set_a | set_b
        return len(intersection) / len(union) if union else 0.0

    @staticmethod
    def _column_name_similarity(name_a: str, name_b: str) -> float:
        set_a = set(name_a)
        set_b = set(name_b)
        if not set_a or not set_b:
            return 0.0
        intersection = set_a & set_b
        if len(intersection) >= 1:
            return len(intersection) / min(len(set_a), len(set_b))
        return 0.0

    def detect_relationships(self, min_confidence: float = 0.15) -> List[Dict]:
        tables = self.get_table_names()
        if len(tables) < 2:
            return []

        all_columns = self.get_all_columns()

        def _is_internal_column(col: str) -> bool:
            skip_prefixes = ["sheet_", "求和项", "计数项", "平均值项", "最大值项", "最小值项"]
            skip_exact = ["sheet"]
            cl = col.strip()
            if cl.lower() in skip_exact:
                return True
            for p in skip_prefixes:
                if cl.lower().startswith(p.lower()):
                    return True
            return False

        candidates = []

        for i in range(len(tables)):
            for j in range(i + 1, len(tables)):
                ta, tb = tables[i], tables[j]
                cols_a, cols_b = all_columns.get(ta, []), all_columns.get(tb, [])

                cols_a = [c for c in cols_a if not _is_internal_column(c)]
                cols_b = [c for c in cols_b if not _is_internal_column(c)]

                for ca in cols_a:
                    values_a = self._sample_column_values(ta, ca)
                    if len(values_a) < 2:
                        continue
                    for cb in cols_b:
                        values_b = self._sample_column_values(tb, cb)
                        if len(values_b) < 2:
                            continue

                        name_sim = self._column_name_similarity(ca, cb)
                        value_overlap = self._column_value_overlap(values_a, values_b)
                        overlap_count = len(set(str(v).strip() for v in values_a) & set(str(v).strip() for v in values_b))

                        is_identical_name = ca.strip() == cb.strip()
                        confidence = value_overlap * 0.7 + name_sim * 0.3

                        if is_identical_name:
                            if overlap_count < 1 and value_overlap < 0.01:
                                continue
                            confidence = max(confidence, 0.85)
                            relation_type = "same_column"
                        elif value_overlap > 0.03 and overlap_count >= 2:
                            relation_type = "value_overlap"
                        elif name_sim > 0.4:
                            relation_type = "name_similarity"
                        else:
                            continue

                        if confidence >= min_confidence:
                            candidates.append({
                                "table_a": ta, "column_a": ca,
                                "table_b": tb, "column_b": cb,
                                "confidence": round(confidence, 3),
                                "relation_type": relation_type,
                                "name_similarity": round(name_sim, 3),
                                "value_overlap": round(value_overlap, 3),
                                "sample_a": values_a[:3],
                                "sample_b": values_b[:3],
                            })

        candidates.sort(key=lambda r: r["confidence"], reverse=True)

        with self._connect() as conn:
            conn.execute("DELETE FROM _relationships")
            for c in candidates[:50]:
                conn.execute(
                    "INSERT INTO _relationships (table_a, column_a, table_b, column_b, confidence, relation_type) VALUES (?, ?, ?, ?, ?, ?)",
                    (c["table_a"], c["column_a"], c["table_b"], c["column_b"], c["confidence"], c["relation_type"]),
                )

        return candidates

    def get_relationships(self) -> List[Dict]:
        with self._connect() as conn:
            rows = conn.execute("SELECT * FROM _relationships ORDER BY confidence DESC").fetchall()
        if not rows:
            try:
                return self.detect_relationships()
            except Exception:
                return []
        return [dict(r) for r in rows]